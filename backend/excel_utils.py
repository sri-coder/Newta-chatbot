"""
excel_utils.py – Lead persistence layer.

Stores leads in a local CSV file (leads.csv).
After every new lead, also writes/refreshes newta_leads.xlsx in the
same data/ folder — same behaviour as the original excel.js auto-download,
but saved to disk instead of the browser.

Files written:
  backend/data/leads.csv          ← append-only raw store
  backend/data/newta_leads.xlsx   ← always up-to-date workbook
"""

from __future__ import annotations
import csv
import io
import os
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Config ────────────────────────────────────────────────────────

DATA_DIR   = Path(os.getenv("DATA_DIR", Path(__file__).parent / "data"))
LEADS_CSV  = DATA_DIR / "leads.csv"
LEADS_XLSX = DATA_DIR / "newta_leads.xlsx"

COLUMNS = [
    "timestamp", "name", "email", "phone", "company",
    "jobTitle", "serviceNeeded", "dataVolume", "sourceSystem",
    "targetSystem", "timeline", "budgetRange", "notes",
]

HEADER_LABELS = [
    "Timestamp", "Name", "Email", "Phone", "Company",
    "Job Title", "Service Needed", "Data Volume", "Source System",
    "Target System", "Timeline", "Budget Range", "Additional Notes",
]

# ── Internal helpers ──────────────────────────────────────────────

def _ensure_data_dir() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def _ensure_csv() -> None:
    _ensure_data_dir()
    if not LEADS_CSV.exists():
        with open(LEADS_CSV, "w", newline="", encoding="utf-8") as f:
            csv.DictWriter(f, fieldnames=COLUMNS).writeheader()


def _build_workbook(leads: list[dict]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_fill = PatternFill("solid", fgColor="006E78")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    center      = Alignment(horizontal="center", vertical="center")

    for col_idx, label in enumerate(HEADER_LABELS, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=label)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center

    for row_idx, lead in enumerate(leads, start=2):
        for col_idx, key in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=lead.get(key, ""))

    for col_idx, label in enumerate(HEADER_LABELS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(label) + 4, 18)

    ws.freeze_panes = "A2"
    return wb


def _write_excel_to_disk() -> None:
    leads = get_all_leads()
    _build_workbook(leads).save(LEADS_XLSX)
    print(f"[LeadStore] Excel updated -> {LEADS_XLSX}  ({len(leads)} row(s))")


# ── Public API ────────────────────────────────────────────────────

def save_lead(lead: dict) -> None:
    """Append a lead to CSV, then refresh newta_leads.xlsx on disk."""
    _ensure_csv()
    row = {col: lead.get(col, "") for col in COLUMNS}
    row["timestamp"] = datetime.now(timezone.utc).isoformat()

    with open(LEADS_CSV, "a", newline="", encoding="utf-8") as f:
        csv.DictWriter(f, fieldnames=COLUMNS).writerow(row)

    _write_excel_to_disk()


def get_all_leads() -> list[dict]:
    _ensure_csv()
    with open(LEADS_CSV, "r", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def export_excel() -> bytes:
    """Return workbook bytes for the /api/leads/export download endpoint."""
    buf = io.BytesIO()
    _build_workbook(get_all_leads()).save(buf)
    buf.seek(0)
    return buf.read()